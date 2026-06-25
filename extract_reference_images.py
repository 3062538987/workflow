import os
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from PIL import Image
import io
import re

# ===== 配置 =====
EXCEL_FILE = "提示词.xlsx"
SHEET_NAME = "Prompt生成表--卖点 原"
OUTPUT_DIR = "reference_images"

def clean_filename(label):
    """清理非法字符，使字符串可用作文件名"""
    if not label:
        return "untitled"
    label = re.sub(r'[\\/*?:"<>|]', '_', label)
    label = re.sub(r'\s+', '_', label)
    return label.strip()

def export_reference_images():
    """从Excel导出参考图，用卖点变量标签命名"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    print("=" * 60)
    print("📸 从 Excel 导出参考图")
    print("=" * 60)
    
    # 加载Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    # 使用 openpyxl_image_loader 加载图片
    try:
        image_loader = SheetImageLoader(ws)
    except Exception as e:
        print(f"⚠️ 无法初始化图片加载器: {e}")
        print("   尝试使用备用方法...")
        return export_reference_images_fallback(wb, ws)
    
    # 获取所有图片
    images = []
    for idx, img in enumerate(ws._images):
        try:
            # 获取图片锚点位置（确定属于哪一行）
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                row = img.anchor._from.row + 1  # openpyxl 从0开始，+1得到实际行号
                col = img.anchor._from.col + 1
                print(f"   📍 图片 {idx+1}: 位于行 {row}, 列 {col}")
            else:
                # 如果无法确定位置，用索引
                row = idx + 2  # 假设从第2行开始（第1行是表头）
                print(f"   ⚠️ 图片 {idx+1}: 无法确定位置，使用行 {row}")
            
            # 读取该行的"卖点变量标签"
            label_cell = ws.cell(row=row, column=16)  # P列（第16列）
            label = label_cell.value if label_cell.value else f"row_{row}"
            
            # 清理文件名
            safe_label = clean_filename(str(label))
            
            # 获取图片数据
            img_data = img._data()
            image = Image.open(io.BytesIO(img_data))
            
            # 保存图片
            filename = f"{safe_label}.png"
            filepath = os.path.join(OUTPUT_DIR, filename)
            
            # 检查是否重名，如果有则加序号
            counter = 1
            base_filename = filename
            while os.path.exists(filepath):
                name_without_ext = safe_label
                filepath = os.path.join(OUTPUT_DIR, f"{name_without_ext}_{counter}.png")
                counter += 1
            
            image.save(filepath, "PNG")
            print(f"   ✅ 已保存: {filename} (标签: {label})")
            images.append(filepath)
            
        except Exception as e:
            print(f"   ❌ 导出第 {idx+1} 张图片失败: {e}")
    
    print(f"\n📊 共导出 {len(images)} 张参考图")
    print(f"📁 保存位置: {OUTPUT_DIR}/")
    return images

def export_reference_images_fallback(wb, ws):
    """备用方法：直接遍历图片"""
    print("\n🔄 使用备用方法导出图片...")
    
    images = []
    for idx, img in enumerate(ws._images):
        try:
            img_data = img._data()
            image = Image.open(io.BytesIO(img_data))
            
            # 用索引命名
            filename = f"ref_{idx+1:03d}.png"
            filepath = os.path.join(OUTPUT_DIR, filename)
            image.save(filepath, "PNG")
            print(f"   ✅ 已保存: {filename}")
            images.append(filepath)
        except Exception as e:
            print(f"   ❌ 导出第 {idx+1} 张图片失败: {e}")
    
    print(f"\n📊 共导出 {len(images)} 张参考图")
    print("⚠️ 图片使用了默认命名 (ref_001.png)，请手动对照重命名")
    return images

if __name__ == "__main__":
    export_reference_images()